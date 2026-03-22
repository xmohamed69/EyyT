# debug_gaze.py
import cv2
import mediapipe as mp
import numpy as np
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from mediapipe.tasks import python as mp_tasks
from mediapipe.tasks.python import vision as mp_vision

# Resolve model path relative to this script so it works
# regardless of which directory you run the script from.
import os as _os
_HERE  = _os.path.dirname(_os.path.abspath(__file__))
_ROOT  = _os.path.dirname(_HERE)   # one level up from debug/
MODEL  = _os.path.join(_ROOT, "data", "face_landmarker.task")

# If still not found, try the current working directory as a fallback
if not _os.path.exists(MODEL):
    MODEL = _os.path.join(_os.getcwd(), "data", "face_landmarker.task")

base = mp_tasks.BaseOptions(model_asset_path=MODEL)
opts = mp_vision.FaceLandmarkerOptions(base_options=base, num_faces=1)
lmk  = mp_vision.FaceLandmarker.create_from_options(opts)

# auto-find camera
cap = None
for i in range(6):
    for backend in (cv2.CAP_ANY, cv2.CAP_DSHOW):
        c = cv2.VideoCapture(i, backend)
        if c.isOpened():
            for _ in range(5):
                ok, f = c.read()
                if ok and f is not None and f.mean() > 1.0:
                    cap = c
                    print(f"Camera found: index={i} backend={backend}")
                    break
        if cap: break
    if cap: break

if not cap:
    print("No camera found"); exit()

for _ in range(10): cap.read()

# ── Excel setup ──────────────────────────────────────────────────────────────
gaze_rows = []   # collect rows in memory; write to Excel on exit
start_ts  = time.time()

# ── Main loop ─────────────────────────────────────────────────────────────────
while True:
    ok, bgr = cap.read()
    if not ok: break
    bgr = cv2.flip(bgr, 1)
    fh, fw = bgr.shape[:2]
    rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
    res = lmk.detect(mp.Image(image_format=mp.ImageFormat.SRGB,
                               data=np.ascontiguousarray(rgb)))
    if res.face_landmarks:
        lms = res.face_landmarks[0]
        li = lms[468]; ri = lms[473]
        lo = lms[33];  li2 = lms[133]
        ro = lms[362]; ri2 = lms[263]
        lt = lms[159]; lb  = lms[145]
        rt = lms[386]; rb  = lms[374]

        lew = li2.x - lo.x;  leh = lb.y - lt.y
        rew = ri2.x - ro.x;  reh = rb.y - rt.y

        lnx = (li.x - lo.x) / lew if abs(lew) > 1e-6 else 0.5
        lny = (li.y - lt.y) / leh if abs(leh) > 1e-6 else 0.5
        rnx = (ri.x - ro.x) / rew if abs(rew) > 1e-6 else 0.5
        rny = (ri.y - rt.y) / reh if abs(reh) > 1e-6 else 0.5

        ex = (lnx + rnx) / 2
        ey = (lny + rny) / 2

        elapsed = round(time.time() - start_ts, 3)
        gaze_rows.append((elapsed, round(ex,4), round(ey,4),
                          round(lnx,4), round(lny,4),
                          round(rnx,4), round(rny,4),
                          round(lew,4), round(leh,4),
                          round(rew,4), round(reh,4)))

        for px,py in ((int(li.x*fw),int(li.y*fh)),(int(ri.x*fw),int(ri.y*fh))):
            cv2.circle(bgr,(px,py),3,(0,255,0),-1)

        cv2.putText(bgr, f"eye_x={ex:.3f}  eye_y={ey:.3f}", (10,30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,255,255), 2)
        cv2.putText(bgr, f"L:{lnx:.2f},{lny:.2f}  R:{rnx:.2f},{rny:.2f}",
                    (10,60), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (200,200,200), 1)
        cv2.putText(bgr, "Look LEFT/RIGHT/UP/DOWN - check ranges",
                    (10, fh-20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (100,200,100), 1)
    else:
        cv2.putText(bgr, "No face", (10,30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,0,255), 2)

    cv2.imshow("Gaze Debug - press Q to quit", bgr)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

# ── Write Excel file ──────────────────────────────────────────────────────────
if gaze_rows:
    wb = Workbook()
    ws = wb.active
    ws.title = "Gaze Data"

    headers = [
        "Time (s)", "Eye X (avg)", "Eye Y (avg)",
        "Left Eye X", "Left Eye Y",
        "Right Eye X", "Right Eye Y",
        "Left Eye Width", "Left Eye Height",
        "Right Eye Width", "Right Eye Height"
    ]

    # header row styling
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # data rows with alternating fill
    light_fill = PatternFill("solid", start_color="D6E4F0")
    for row_idx, row in enumerate(gaze_rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            if row_idx % 2 == 0:
                cell.fill = light_fill

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    n = len(gaze_rows)
    summary_headers = ["Metric", "Eye X (avg)", "Eye Y (avg)", "Left Eye X", "Left Eye Y", "Right Eye X", "Right Eye Y"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    col_letters = ["B", "C", "D", "E", "F", "G"]  # columns 2-7 in Gaze Data sheet
    metrics = ["Min", "Max", "Average", "Std Dev"]
    formulas = ["MIN", "MAX", "AVERAGE", "STDEV"]
    for r, (metric, func) in enumerate(zip(metrics, formulas), 2):
        ws2.cell(row=r, column=1, value=metric).font = Font(bold=True, name="Arial", size=10)
        for c, col_letter in enumerate(col_letters, 2):
            ws2.cell(row=r, column=c,
                     value=f"='Gaze Data'!{col_letter}2:{col_letter}{n+1}").value = \
                f"={func}('Gaze Data'!{col_letter}2:{col_letter}{n+1})"
            ws2.cell(row=r, column=c).number_format = "0.0000"
            ws2.cell(row=r, column=c).font = Font(name="Arial", size=10)

    ws2.cell(row=6, column=1, value="Total Frames").font = Font(bold=True, name="Arial", size=10)
    ws2.cell(row=6, column=2, value=n).font = Font(name="Arial", size=10)
    ws2.cell(row=7, column=1, value="Session Duration (s)").font = Font(bold=True, name="Arial", size=10)
    ws2.cell(row=7, column=2, value=round(time.time() - start_ts, 2)).font = Font(name="Arial", size=10)
    ws2.cell(row=8, column=1, value="Recorded At").font = Font(bold=True, name="Arial", size=10)
    ws2.cell(row=8, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = Font(name="Arial", size=10)

    # auto-size columns in both sheets
    for sheet in (ws, ws2):
        for col in sheet.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    import os
    docs_folder = os.path.join(os.path.expanduser("~"), "Documents")
    os.makedirs(docs_folder, exist_ok=True)
    filename = os.path.join(docs_folder, f"gaze_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filename)
    print(f"Saved {n} frames → {filename}")
else:
    print("No gaze data recorded.")